#!/usr/bin/env python3
"""检索回归大测试 —— 专业指标台账导出器。

设计目标（按学术论文式呈现）：
  * 运行台账为「纵向」布局：指标做行、每一次大测试做一列，便于逐列对比版本差异。
  * 开发期预热测试与「正式测试」用分割线隔开：从首个 100 用例新测试集那次起算正式测试。
  * 折线图暂不绘制（等算法优化后再重测一轮再出图：横轴=第几次正式测试，纵轴=精度，
    展示每次更新带来的变化）。
  * "基线 vs 最新" 同配置对比表，直接给出每个核心指标的变化量 Δ。
  * 重跑脚本即自动并入新产生的 report.json（数据自动追加）。
  * 人工"备注"按 run_id 回填，重跑后保留，方便标注 "baseline / +破笼 / +gating" 等。

用法：
    python scripts/export-retrieval-regression-excel.py
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
INK = "1F3A5F"          # 深蓝：标题 / 表头底色（其上文字一律白色）
PREFORMAL = "7F7F7F"    # 灰：预热测试列头
DIVIDER = "C00000"      # 红：正式测试分割线
ZEBRA = "F2F6FB"        # 浅蓝：斑马行
GRID = "BFBFBF"
WHITE = "FFFFFF"

# 首个达到该用例数的测试，视为「正式测试」起点（新测试集为 100 用例）。
FORMAL_MIN_CASES = 90

# ---- 纵向台账的字段行（名称, summary/派生键, 数字格式）----
# _local/_health/_note/_run_id/_pass_rate 为派生键；其余取自 summary 或顶层。
LEDGER_FIELDS = [
    ("日期时间", "_local", "yyyy-mm-dd hh:mm"),
    ("模式", "evaluation_mode", None),
    ("Profile", "profile", None),
    ("用例数", "case_count", "0"),
    ("Top-1 文档命中", "top1_document_hit_rate", "0.0%"),
    ("Top-3 文档召回", "top3_document_recall_rate", "0.0%"),
    ("Top-1 Chunk 命中", "top1_chunk_hit_rate", "0.0%"),
    ("Top-5 Chunk 召回", "top5_chunk_recall_rate", "0.0%"),
    ("Chunk MRR", "chunk_mrr", "0.000"),
    ("Citation 有效率", "citation_validity_rate", "0.0%"),
    ("拒答正确率", "reject_correctness_rate", "0.0%"),
    ("综合通过率", "_pass_rate", "0.0%"),
    ("Rerank 应用率", "rerank_applied_rate", "0.0%"),
    ("服务/重排", "_health", None),
    ("备注", "_note", None),
    ("run_id", "_run_id", None),  # 末行，回填备注用，导出后隐藏
]

CASE_COLS = [
    ("run_id", "run_id"),
    ("测试标签", "label"),
    ("用例", "id"),
    ("模式", "mode"),
    ("状态", "status"),
    ("通过", "passed"),
    ("问题", "query"),
    ("文档命中名次", "document_hit_rank"),
    ("Chunk 命中名次", "chunk_hit_rank"),
    ("Top-3 文档召回", "top3_document_recall"),
    ("Top-5 Chunk 召回", "top5_chunk_recall"),
    ("Citation 有效", "citation_valid"),
    ("拒答正确", "reject_correct"),
    ("门控决策", "gating_decision_reason"),
]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--reports-root", default="target/retrieval-regression")
    p.add_argument("--output", default="docs/qa/retrieval_regression_metrics.xlsx")
    return p.parse_args()


def to_local(raw: Any) -> datetime | None:
    try:
        ts = int(str(raw))
    except (TypeError, ValueError):
        return None
    return datetime.fromtimestamp(ts).replace(tzinfo=None)


def case_passed(case: dict[str, Any]) -> bool:
    if case.get("timed_out"):
        return False
    if case.get("mode") == "refuse":
        return bool(case.get("reject_correct"))
    return bool(
        case.get("top3_document_recall")
        and case.get("top5_chunk_recall")
        and case.get("citation_valid")
    )


def pass_rate(report: dict[str, Any]) -> float | None:
    cases = report.get("cases", [])
    if not cases:
        return None
    return sum(1 for c in cases if case_passed(c)) / len(cases)


def load_reports(root: Path) -> list[dict[str, Any]]:
    reports = []
    for path in root.rglob("report.json"):
        try:
            report = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        report["_report_json"] = str(path.resolve())
        report["_run_id"] = path.parent.name
        reports.append(report)
    reports.sort(key=lambda r: str(r.get("generated_at_utc", "")))
    return reports


def formal_start_index(valid: list[dict[str, Any]]) -> int | None:
    """返回首个「正式测试」在 valid 中的下标；之后（按时间）全部视为正式测试。"""
    for i, r in enumerate(valid):
        if r.get("summary", {}).get("case_count", 0) >= FORMAL_MIN_CASES:
            return i
    return None


def load_existing_notes(output_path: Path) -> dict[str, str]:
    """从已有（纵向）台账回收人工备注（按 run_id），重跑后不丢失。"""
    if not output_path.exists():
        return {}
    try:
        wb = load_workbook(output_path, read_only=True, data_only=True)
    except Exception:
        return {}
    if "运行台账" not in wb.sheetnames:
        return {}
    ws = wb["运行台账"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    # 纵向布局：找到 A 列标签为 "run_id" / "备注" 的两行，按列对位。
    run_row = note_row = None
    for row in rows:
        if not row:
            continue
        label = row[0]
        if label == "run_id":
            run_row = row
        elif label == "备注":
            note_row = row
    notes: dict[str, str] = {}
    if run_row and note_row:
        for c in range(1, len(run_row)):
            rid = run_row[c]
            note = note_row[c] if c < len(note_row) else None
            if rid and note:
                notes[str(rid)] = str(note)
    return notes


# --------------------------- styling helpers ---------------------------

def title_cell(ws, coord: str, text: str, size: int = 16, fill: str = INK) -> None:
    """统一标题样式：彩色底 + 白字（杜绝黑字看不见）。"""
    cell = ws[coord]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=FONT_NAME, size=size, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def thin_border(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    side = Side(style="thin", color=GRID)
    border = Border(left=side, right=side, top=side, bottom=side)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


def header_row_style(ws, row: int, ncol: int) -> None:
    fill = PatternFill("solid", fgColor=INK)
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# --------------------------- sheets ---------------------------

def build_ledger(wb: Workbook, valid: list[dict[str, Any]], notes: dict[str, str]):
    """纵向台账：第 1 列 = 指标名；其后每一列 = 一次大测试；中间用分割线列隔开预热/正式。"""
    ws = wb.create_sheet("运行台账")
    fstart = formal_start_index(valid)

    # 列计划：run / divider。divider 仅在「既有预热又有正式」时插入一次。
    col_plan: list[tuple] = []
    for i, r in enumerate(valid):
        if fstart is not None and i == fstart and i > 0:
            col_plan.append(("divider",))
        col_plan.append(("run", r, i + 1, (fstart is not None and i >= fstart)))

    ncol = 1 + len(col_plan)

    # ---- 表头行（row 1）----
    ws.cell(row=1, column=1, value="指标 ＼ 测试")
    for idx, item in enumerate(col_plan, start=2):
        if item[0] == "divider":
            ws.cell(row=1, column=idx, value="正式 →")
        else:
            _, r, seq, _is_formal = item
            mode = r.get("evaluation_mode", "")
            short = {"offline_deterministic": "offline", "live_embedding": "live"}.get(mode, mode)
            ws.cell(row=1, column=idx, value=f"#{seq}\n{short}·{r.get('profile','')}")
    header_row_style(ws, 1, ncol)

    # ---- 字段行 ----
    short_health = lambda r: f"{r.get('service_health','')}/{r.get('rerank_health','')}"
    for ri, (fname, key, fmt) in enumerate(LEDGER_FIELDS, start=2):
        # A 列：指标名（深蓝底白字，属"标题")
        a = ws.cell(row=ri, column=1, value=fname)
        a.fill = PatternFill("solid", fgColor=INK)
        a.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for idx, item in enumerate(col_plan, start=2):
            cell = ws.cell(row=ri, column=idx)
            if item[0] == "divider":
                continue
            _, r, _seq, _is_formal = item
            summary = r.get("summary", {})
            run_id = r.get("_run_id", "")
            if key == "_local":
                val = to_local(r.get("generated_at_utc"))
            elif key == "_health":
                val = short_health(r)
            elif key == "_note":
                val = notes.get(run_id, "")
            elif key == "_run_id":
                val = run_id
            elif key == "_pass_rate":
                val = pass_rate(r)
            elif key in {"evaluation_mode", "profile"}:
                val = r.get(key, "")
            elif key == "case_count":
                val = summary.get("case_count", 0)
            else:
                val = summary.get(key)
            cell.value = val
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal="center", vertical="center")

    nrow = 1 + len(LEDGER_FIELDS)

    # ---- 列样式：分割线列 + 预热/正式列头配色 + 斑马行 ----
    for idx, item in enumerate(col_plan, start=2):
        letter = get_column_letter(idx)
        if item[0] == "divider":
            ws.column_dimensions[letter].width = 4
            for r in range(1, nrow + 1):
                c = ws.cell(row=r, column=idx)
                c.fill = PatternFill("solid", fgColor=DIVIDER)
            top = ws.cell(row=1, column=idx)
            top.font = Font(name=FONT_NAME, size=9, bold=True, color=WHITE)
            top.alignment = Alignment(horizontal="center", vertical="center",
                                      wrap_text=True, text_rotation=90)
            continue
        ws.column_dimensions[letter].width = 14
        # 列头配色：正式=深蓝，预热=灰（均白字）
        is_formal = item[3]
        ws.cell(row=1, column=idx).fill = PatternFill(
            "solid", fgColor=(INK if is_formal else PREFORMAL))

    # 斑马行（隔行浅蓝），不覆盖 A 列与分割线列
    divider_cols = {idx for idx, item in enumerate(col_plan, start=2) if item[0] == "divider"}
    for r in range(2, nrow + 1):
        if r % 2 == 1:
            for idx in range(2, ncol + 1):
                if idx in divider_cols:
                    continue
                if ws.cell(row=r, column=idx).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=r, column=idx).fill = PatternFill("solid", fgColor=ZEBRA)

    thin_border(ws, 1, 1, nrow, ncol)
    ws.column_dimensions["A"].width = 18
    ws.row_dimensions[1].height = 30
    # run_id 行隐藏
    ws.row_dimensions[nrow].hidden = True
    ws.freeze_panes = "B2"
    return ws


def build_cases(wb: Workbook, valid: list[dict[str, Any]], label_by_run: dict[str, str]):
    ws = wb.create_sheet("用例明细")
    ws.append([h for h, _ in CASE_COLS])
    for report in valid:
        rid = report.get("_run_id", "")
        for case in report.get("cases", []):
            vals = {
                "run_id": rid,
                "label": label_by_run.get(rid, ""),
                "id": case.get("id", ""),
                "mode": case.get("mode", ""),
                "status": case.get("status", ""),
                "passed": case_passed(case),
                "query": case.get("query", ""),
                "document_hit_rank": case.get("document_hit_rank"),
                "chunk_hit_rank": case.get("chunk_hit_rank"),
                "top3_document_recall": case.get("top3_document_recall"),
                "top5_chunk_recall": case.get("top5_chunk_recall"),
                "citation_valid": case.get("citation_valid"),
                "reject_correct": case.get("reject_correct"),
                "gating_decision_reason": case.get("gating_decision_reason", ""),
            }
            ws.append([vals[k] for _, k in CASE_COLS])
    ncol = len(CASE_COLS)
    header_row_style(ws, 1, ncol)
    ws.freeze_panes = "C2"
    for i, (h, _) in enumerate(CASE_COLS, start=1):
        w = 40 if h == "问题" else (36 if h == "run_id" else 14)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["A"].hidden = True
    return ws


def build_failures(wb: Workbook, valid: list[dict[str, Any]], label_by_run: dict[str, str]):
    ws = wb.create_sheet("失败用例")
    ws.append([h for h, _ in CASE_COLS])
    for report in valid:
        rid = report.get("_run_id", "")
        for case in report.get("cases", []):
            if case_passed(case):
                continue
            vals = {
                "run_id": rid, "label": label_by_run.get(rid, ""), "id": case.get("id", ""),
                "mode": case.get("mode", ""), "status": case.get("status", ""), "passed": False,
                "query": case.get("query", ""), "document_hit_rank": case.get("document_hit_rank"),
                "chunk_hit_rank": case.get("chunk_hit_rank"),
                "top3_document_recall": case.get("top3_document_recall"),
                "top5_chunk_recall": case.get("top5_chunk_recall"),
                "citation_valid": case.get("citation_valid"),
                "reject_correct": case.get("reject_correct"),
                "gating_decision_reason": case.get("gating_decision_reason", ""),
            }
            ws.append([vals[k] for _, k in CASE_COLS])
    ncol = len(CASE_COLS)
    header_row_style(ws, 1, ncol)
    ws.freeze_panes = "C2"
    for i, (h, _) in enumerate(CASE_COLS, start=1):
        w = 40 if h == "问题" else (36 if h == "run_id" else 14)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["A"].hidden = True


def build_raw(wb: Workbook, all_reports: list[dict[str, Any]]):
    ws = wb.create_sheet("原始字段")
    headers = ["run_id", "时间(本地)", "模式", "Profile", "用例数", "service_health",
               "rerank_health", "case_timeout", "index_prep_ms", "preparation_error",
               "watch_root", "db_path", "report_json"]
    ws.append(headers)
    for r in all_reports:
        s = r.get("summary", {})
        ws.append([
            r.get("_run_id", ""), to_local(r.get("generated_at_utc")),
            r.get("evaluation_mode", ""), r.get("profile", ""), s.get("case_count", 0),
            r.get("service_health", ""), r.get("rerank_health", ""),
            r.get("case_timeout_count", 0), r.get("index_prep_ms"),
            r.get("preparation_error") or "", r.get("watch_root", ""),
            r.get("db_path", ""), r.get("_report_json", ""),
        ])
    header_row_style(ws, 1, len(headers))
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 18 if i < 9 else 40
    ws.column_dimensions["B"].number_format = "yyyy-mm-dd hh:mm"
    ws.freeze_panes = "A2"


def build_dashboard(wb: Workbook, valid: list[dict[str, Any]]):
    """趋势页：折线图暂缓（待优化后重测再出图）；先放同配置 基线 vs 最新 对比表。"""
    ws = wb.create_sheet("核心趋势", 0)
    ws.sheet_view.showGridLines = False
    title_cell(ws, "A1", "检索回归大测试 · 核心指标", size=16)
    for c in range(2, 5):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=INK)

    ws["A2"] = ("折线图暂缓：等算法优化后再正式重测一轮再绘制——"
                "横轴=第几次正式测试，纵轴=精度，展示每次更新带来的变化。"
                "当前先看下方「同配置对比」与「运行台账（纵向）」。")
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="606060")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:D3")

    # ---- 同配置 基线 vs 最新 对比表 ----
    cmp_top = 5
    title_cell(ws, f"A{cmp_top}", "同配置 基线 vs 最新 对比", size=13)
    for c in range(2, 5):
        ws.cell(row=cmp_top, column=c).fill = PatternFill("solid", fgColor=INK)

    group: list[dict[str, Any]] = []
    latest = valid[-1] if valid else None
    if latest:
        sig = (latest.get("evaluation_mode"), latest.get("profile"),
               latest.get("summary", {}).get("case_count"))
        group = [r for r in valid
                 if (r.get("evaluation_mode"), r.get("profile"),
                     r.get("summary", {}).get("case_count")) == sig]

    if latest and len(group) >= 2:
        base_run, last_run = group[0], group[-1]
        base_seq = valid.index(base_run) + 1
        last_seq = valid.index(last_run) + 1
        ws.cell(row=cmp_top + 1, column=1,
                value=(f"配置：{latest.get('evaluation_mode')}·{latest.get('profile')}·"
                       f"{sig[2]}用例    基线=#{base_seq}    最新=#{last_seq}"
                       f"    （该配置共 {len(group)} 次）")).font = \
            Font(name=FONT_NAME, size=9, italic=True, color="606060")
        head = ["核心指标", "基线", "最新", "变化 Δ"]
        hr = cmp_top + 2
        header_row_style(ws, hr, len(head))
        for c, h in enumerate(head, start=1):
            ws.cell(row=hr, column=c, value=h)
        header_row_style(ws, hr, len(head))
        metric_keys = {
            "Top-1 文档命中": "top1_document_hit_rate",
            "Top-3 文档召回": "top3_document_recall_rate",
            "Top-1 Chunk 命中": "top1_chunk_hit_rate",
            "Top-5 Chunk 召回": "top5_chunk_recall_rate",
            "Chunk MRR": "chunk_mrr",
            "Citation 有效率": "citation_validity_rate",
            "拒答正确率": "reject_correctness_rate",
            "综合通过率": None,
            "Rerank 应用率": "rerank_applied_rate",
        }
        base_sum, last_sum = base_run.get("summary", {}), last_run.get("summary", {})

        def metric_value(report, summary, key):
            if key is None:
                return pass_rate(report)
            return summary.get(key)

        for j, (name, key) in enumerate(metric_keys.items()):
            r = hr + 1 + j
            fmt = "0.000" if name == "Chunk MRR" else "0.0%"
            bval = metric_value(base_run, base_sum, key)
            lval = metric_value(last_run, last_sum, key)
            dval = (lval - bval) if (isinstance(bval, (int, float)) and isinstance(lval, (int, float))) else None
            ws.cell(row=r, column=1, value=name).font = Font(name=FONT_NAME, size=10)
            b = ws.cell(row=r, column=2, value=bval)
            l = ws.cell(row=r, column=3, value=lval)
            d = ws.cell(row=r, column=4, value=dval)
            for cell in (b, l):
                cell.number_format = fmt
                cell.font = Font(name=FONT_NAME, size=10)
                cell.alignment = Alignment(horizontal="center")
            d.number_format = ('[Green]"▲ "0.000;[Red]"▼ "0.000;0.000'
                               if name == "Chunk MRR"
                               else '[Green]"▲ "0.0%;[Red]"▼ "0.0%;0.0%')
            d.font = Font(name=FONT_NAME, size=10, bold=True)
            d.alignment = Alignment(horizontal="center")
        thin_border(ws, hr, 1, hr + len(metric_keys), 4)
    else:
        cfg = (f"{latest.get('evaluation_mode')}·{latest.get('profile')}·"
               f"{latest.get('summary', {}).get('case_count')}用例") if latest else "—"
        ws.cell(row=cmp_top + 1, column=1,
                value=(f"当前最新配置（{cfg}）暂无同配置历史可对比；"
                       f"需 ≥2 次相同 模式/Profile/用例数 的测试才能给出 Δ。")).font = \
            Font(name=FONT_NAME, size=10, italic=True, color="606060")

    for col, w in {"A": 22, "B": 14, "C": 14, "D": 16}.items():
        ws.column_dimensions[col].width = w


def build_notes(wb: Workbook, reports_root: Path, output_path: Path, n_valid: int, n_total: int,
                fstart: int | None):
    ws = wb.create_sheet("说明")
    ws.sheet_view.showGridLines = False
    title_cell(ws, "A1", "说明", size=14)
    ws.cell(row=1, column=2).fill = PatternFill("solid", fgColor=INK)
    formal_note = ("从第 %d 次起为正式测试（首个 ≥%d 用例的新测试集），台账中以红色分割线列隔开；"
                   "之前为开发期预热测试。" % (fstart + 1, FORMAL_MIN_CASES)) if fstart is not None \
        else "尚无达到 %d 用例的正式测试。" % FORMAL_MIN_CASES
    rows = [
        ("用途", "记录每一次检索回归大测试；正式测试积累后再绘制精度随版本变化的折线图。"),
        ("台账布局", "「运行台账」为纵向：指标做行、每次测试做列，逐列对比版本差异。"),
        ("正式测试", formal_note),
        ("刷新方式", "跑完回归后执行：python scripts/export-retrieval-regression-excel.py"),
        ("自动追加", "脚本扫描 report.json 全量重建；新测试重跑即并入，无需手工录入。"),
        ("备注保留", "纵向台账「备注」行按 run_id 回填，重跑不丢失。可标注 baseline / +破笼 / +gating 等。"),
        ("有效口径", "趋势与对比仅含 case_count>0 的测试；准备失败/0 用例的运行见「原始字段」。"),
        ("通过判定", "answer：Top-3 文档召回 ∧ Top-5 Chunk 召回 ∧ Citation 有效；refuse：拒答正确。"),
        ("折线计划", "横轴=第几次正式测试，纵轴=各精度指标；每次算法更新后重测一轮，观察折线变化。"),
        ("报告来源", str(reports_root.resolve())),
        ("输出文件", str(output_path.resolve())),
        ("本次纳入", f"有效测试 {n_valid} 次（共发现报告 {n_total} 份）。"),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(name=FONT_NAME, size=10, bold=True, color=INK)
        c = ws.cell(row=i, column=2, value=v)
        c.font = Font(name=FONT_NAME, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 96


def main() -> None:
    args = parse_args()
    reports_root = Path(args.reports_root)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    all_reports = load_reports(reports_root)
    valid = [r for r in all_reports if r.get("summary", {}).get("case_count", 0) > 0]
    notes = load_existing_notes(output_path)
    fstart = formal_start_index(valid)

    wb = Workbook()
    wb.remove(wb.active)

    build_ledger(wb, valid, notes)
    label_by_run = {r.get("_run_id", ""): f"#{i} {r.get('evaluation_mode','')}·{r.get('profile','')}"
                    for i, r in enumerate(valid, start=1)}
    build_cases(wb, valid, label_by_run)
    build_failures(wb, valid, label_by_run)
    build_raw(wb, all_reports)
    build_dashboard(wb, valid)
    build_notes(wb, reports_root, output_path, len(valid), len(all_reports), fstart)

    # sheet 顺序：核心趋势 / 运行台账 / 用例明细 / 失败用例 / 原始字段 / 说明
    try:
        wb.save(output_path)
        target = output_path
    except PermissionError:
        target = output_path.with_name(
            f"{output_path.stem}.{datetime.now():%Y%m%d_%H%M%S}{output_path.suffix}")
        wb.save(target)
        print(f"[warn] {output_path} 被占用（可能在 Excel 中打开），已改写到 {target.name}")
    print(f"wrote {target.resolve()} | valid runs: {len(valid)} | "
          f"formal start: #{fstart + 1 if fstart is not None else '—'} | total reports: {len(all_reports)}")


if __name__ == "__main__":
    main()
